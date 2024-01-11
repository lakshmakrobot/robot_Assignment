import openpyxl

def append_rows(file_path, sheet_name, row1 , row2 , row3):
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the sheet or create a new one if it doesn't exist
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    rows = [row1,row2,row3]
    # Append each row to the sheet
    
    sheet.append(rows)

    # Save the changes
    workbook.save(file_path)

if __name__ == "__main__":
    # Specify the file path, sheet name, and rows to append
    file_path = "example.xlsx"
    sheet_name = "Sheet1"
    new_rows = ["NewValue1", "NewValue2", "NewValue3"]
    # Add more rows as needed
    # Append rows to the existing or newly created XLSX file
    append_rows(file_path, sheet_name, new_rows)
